from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for, Response
from werkzeug.middleware.proxy_fix import ProxyFix
import os
import json
import csv
import re
from datetime import datetime, timedelta, date
import qrcode
import barcode
from barcode.writer import ImageWriter
from io import BytesIO, StringIO
import base64
from functools import wraps
import secrets
from dotenv import load_dotenv
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_talisman import Talisman
from flask_wtf.csrf import CSRFProtect
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from supabase import create_client

load_dotenv()

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', secrets.token_hex(32))
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=int(os.getenv('SESSION_TIMEOUT_MINUTES', '30')))
app.config['SESSION_COOKIE_HTTPONLY'] = os.getenv('SESSION_COOKIE_HTTPONLY', 'True') == 'True'
samesite_value = os.getenv('SESSION_COOKIE_SAMESITE', 'Lax')
app.config['SESSION_COOKIE_SAMESITE'] = None if samesite_value == 'None' else samesite_value
app.config['SESSION_COOKIE_SECURE'] = os.getenv('SESSION_COOKIE_SECURE', 'False') == 'True'
app.config['SESSION_COOKIE_NAME'] = 'session'
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024

UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', 'static/uploads/items')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Supabase clients
SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_ANON_KEY = os.getenv('SUPABASE_ANON_KEY')
SUPABASE_SERVICE_KEY = os.getenv('SUPABASE_SERVICE_KEY')

sb = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)   # all DB + admin auth ops
sb_auth = create_client(SUPABASE_URL, SUPABASE_ANON_KEY) # user sign-in

limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=[os.getenv('API_RATE_LIMIT', '100 per minute')],
    storage_uri=os.getenv('RATELIMIT_STORAGE_URL', 'memory://')
)

csrf = CSRFProtect(app)
app.config['WTF_CSRF_CHECK_DEFAULT'] = False
app.config['WTF_CSRF_HEADERS'] = ['X-CSRFToken']

@app.template_filter('fromjson')
def fromjson_filter(value):
    if isinstance(value, str):
        return json.loads(value)
    return value

if os.getenv('SESSION_COOKIE_SECURE', 'False') == 'True':
    Talisman(app,
             force_https=True,
             strict_transport_security=True,
             strict_transport_security_max_age=31536000,
             content_security_policy={
                 'default-src': "'self'",
                 'script-src': "'self' 'unsafe-inline' https://cdn.jsdelivr.net",
                 'style-src': "'self' 'unsafe-inline'",
                 'img-src': "'self' data:",
                 'font-src': "'self'",
                 'connect-src': "'self'",
                 'frame-ancestors': "'none'",
             })
else:
    @app.after_request
    def add_security_headers(response):
        response.headers['X-Content-Type-Options'] = 'nosniff'
        response.headers['X-Frame-Options'] = 'DENY'
        response.headers['X-XSS-Protection'] = '1; mode=block'
        response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
        return response

failed_login_attempts = {}
MAX_LOGIN_ATTEMPTS = int(os.getenv('MAX_LOGIN_ATTEMPTS', '5'))
LOCKOUT_DURATION = timedelta(minutes=15)

# ─── Decorators ───────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        if not session.get('is_admin'):
            return jsonify({'success': False, 'message': 'Admin-Rechte erforderlich'}), 403
        return f(*args, **kwargs)
    return decorated_function

def csrf_protect_api():
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if request.method in ['POST', 'PUT', 'DELETE']:
                csrf.protect()
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# ─── Helpers ──────────────────────────────────────────────────────────────────

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def secure_filename_custom(filename):
    if '.' in filename:
        name, ext = filename.rsplit('.', 1)
        ext = ext.lower()
    else:
        ext = ''
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    random_suffix = secrets.token_hex(4)
    return f"{timestamp}_{random_suffix}.{ext}" if ext else f"{timestamp}_{random_suffix}"

def sanitize_string(value, max_length=255):
    if not value:
        return value
    value = str(value).strip()
    if len(value) > max_length:
        value = value[:max_length]
    return value

def validate_number(value, min_val=None, max_val=None):
    try:
        num = float(value) if '.' in str(value) else int(value)
        if min_val is not None and num < min_val:
            return None
        if max_val is not None and num > max_val:
            return None
        return num
    except (ValueError, TypeError):
        return None

def validate_password(password):
    if len(password) < 8:
        return False, "Passwort muss mindestens 8 Zeichen lang sein"
    if not any(c.isupper() for c in password):
        return False, "Passwort muss mindestens einen Großbuchstaben enthalten"
    if not any(c.islower() for c in password):
        return False, "Passwort muss mindestens einen Kleinbuchstaben enthalten"
    if not any(c.isdigit() for c in password):
        return False, "Passwort muss mindestens eine Zahl enthalten"
    return True, "OK"

def create_slug(text):
    text = text.lower()
    text = text.replace('ä', 'ae').replace('ö', 'oe').replace('ü', 'ue').replace('ß', 'ss')
    text = re.sub(r'[^a-z0-9-]', '-', text)
    text = re.sub(r'-+', '-', text)
    return text.strip('-')

def is_account_locked(ip_address):
    if ip_address in failed_login_attempts:
        data = failed_login_attempts[ip_address]
        if 'locked_until' in data and data['locked_until'] > datetime.now():
            return True, data['locked_until']
    return False, None

def record_failed_login(ip_address):
    if ip_address not in failed_login_attempts:
        failed_login_attempts[ip_address] = {'count': 0}
    failed_login_attempts[ip_address]['count'] += 1
    if failed_login_attempts[ip_address]['count'] >= MAX_LOGIN_ATTEMPTS:
        failed_login_attempts[ip_address]['locked_until'] = datetime.now() + LOCKOUT_DURATION
        return True
    return False

def reset_failed_logins(ip_address):
    if ip_address in failed_login_attempts:
        del failed_login_attempts[ip_address]

def get_base_url():
    return request.host_url.rstrip('/')

def _normalize_item(item):
    """Flatten supabase FK-embedded category/location/group into flat fields."""
    if item is None:
        return item
    if isinstance(item.get('categories'), dict):
        item['category_name'] = item.pop('categories', {}).get('name')
    elif 'categories' in item:
        item['category_name'] = None
        item.pop('categories', None)
    if isinstance(item.get('locations'), dict):
        item['location_name'] = item.pop('locations', {}).get('name')
    elif 'locations' in item:
        item['location_name'] = None
        item.pop('locations', None)
    # self-join for group name comes back as list or dict depending on query
    grp = item.pop('group_item', None)
    if isinstance(grp, dict):
        item['group_name'] = grp.get('name')
    elif isinstance(grp, list) and grp:
        item['group_name'] = grp[0].get('name')
    else:
        item.setdefault('group_name', None)
    return item

# ─── QR / Barcode ─────────────────────────────────────────────────────────────

def generate_qr_code(item_id, item_data):
    qr_data = f"{get_base_url()}/item/{item_id}"
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=10, border=4)
    qr.add_data(qr_data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    return buffer

def generate_qr_code_base64(item_id, item_data):
    buffer = generate_qr_code(item_id, item_data)
    return base64.b64encode(buffer.getvalue()).decode('utf-8')

def generate_barcode(item_id, item_data):
    barcode_value = item_data.get('barcode') or f"ITEM{item_id:08d}"
    try:
        code128 = barcode.get('code128', barcode_value, writer=ImageWriter())
        buffer = BytesIO()
        code128.write(buffer)
        buffer.seek(0)
        return buffer
    except Exception:
        buffer = BytesIO()
        img = qrcode.make(barcode_value)
        img.save(buffer, format='PNG')
        buffer.seek(0)
        return buffer

def generate_barcode_base64(item_id, item_data):
    buffer = generate_barcode(item_id, item_data)
    return base64.b64encode(buffer.getvalue()).decode('utf-8')

# ─── Auth routes ──────────────────────────────────────────────────────────────

@app.route('/landing')
def landing():
    return render_template('landing.html')

@app.route('/register', methods=['GET'])
def register_page():
    if 'logged_in' in session:
        return redirect(url_for('index'))
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit(os.getenv('LOGIN_RATE_LIMIT', '5 per minute'))
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')

        ip_address = get_remote_address()
        is_locked, locked_until = is_account_locked(ip_address)
        if is_locked:
            remaining = int((locked_until - datetime.now()).total_seconds() // 60)
            return render_template('login.html',
                error=f'Konto gesperrt. Bitte warten Sie noch {remaining} Minuten.')

        try:
            auth_resp = sb_auth.auth.sign_in_with_password({"email": email, "password": password})
            auth_user = auth_resp.user
            if not auth_user:
                raise Exception("Login fehlgeschlagen")

            # Fetch user row from users table
            user_rows = sb.table('users').select(
                '*, organizations(name, is_active)'
            ).eq('auth_id', auth_user.id).limit(1).execute().data

            if not user_rows:
                return render_template('login.html', error='Benutzer nicht gefunden')

            user = user_rows[0]
            org = user.get('organizations') or {}
            if not org.get('is_active', True):
                return render_template('login.html', error='Organisation deaktiviert')

            # Update last_login
            sb.table('users').update({'last_login': datetime.now().isoformat()}).eq('id', user['id']).execute()

            reset_failed_logins(ip_address)
            session['logged_in'] = True
            session['user_id'] = user['id']
            session['auth_id'] = str(auth_user.id)
            session['username'] = user['username']
            session['email'] = email
            session['organization_id'] = user['organization_id']
            session['organization_name'] = org.get('name', '')
            session['is_admin'] = bool(user.get('is_admin'))
            session['is_org_owner'] = bool(user.get('is_org_owner'))

            if request.form.get('remember'):
                session.permanent = True
            else:
                session.permanent = False

            return redirect(url_for('index'))

        except Exception as e:
            just_locked = record_failed_login(ip_address)
            if just_locked:
                return render_template('login.html',
                    error=f'Konto gesperrt! Zu viele fehlgeschlagene Versuche. Bitte in {int(LOCKOUT_DURATION.total_seconds() // 60)} Minuten erneut versuchen.')
            attempts_left = MAX_LOGIN_ATTEMPTS - failed_login_attempts.get(ip_address, {}).get('count', 0)
            return render_template('login.html',
                error=f'Falsche E-Mail oder falsches Passwort. Noch {attempts_left} Versuche übrig.')

    if 'logged_in' in session:
        return redirect(url_for('index'))

    success = None
    if request.args.get('registered') == 'true':
        success = 'Organisation erfolgreich erstellt! Bitte melden Sie sich an.'
    return render_template('login.html', success=success)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ─── Page routes ──────────────────────────────────────────────────────────────

@app.route('/')
def index():
    if 'logged_in' in session:
        return render_template('index.html',
                               username=session.get('username'),
                               organization_name=session.get('organization_name'),
                               is_admin=session.get('is_admin'),
                               is_org_owner=session.get('is_org_owner'))
    return render_template('landing.html')

@app.route('/profile')
@login_required
def profile():
    user_rows = sb.table('users').select(
        'email, notify_low_stock, notify_maintenance'
    ).eq('id', session.get('user_id')).limit(1).execute().data
    user = user_rows[0] if user_rows else {}
    return render_template('profile.html',
                           username=session.get('username'),
                           is_admin=session.get('is_admin'),
                           email=session.get('email', user.get('email', '')),
                           notify_low_stock=user.get('notify_low_stock', True),
                           notify_maintenance=user.get('notify_maintenance', True))

@app.route('/settings')
@login_required
def settings():
    return render_template('settings.html',
                           username=session.get('username'),
                           is_admin=session.get('is_admin'))

@app.route('/users')
@admin_required
def users_page():
    return render_template('users.html')

@app.route('/offline')
def offline():
    return render_template('offline.html')

@app.route('/impressum')
def impressum():
    return render_template('impressum.html')

@app.route('/datenschutz')
def datenschutz():
    return render_template('datenschutz.html')

@app.route('/agb')
def agb():
    return render_template('agb.html')

@app.route('/label-designer')
@login_required
def label_designer():
    return render_template('label_designer.html')

@app.route('/label-preview')
@login_required
def label_preview():
    return render_template('label_preview.html')

@app.route('/item/<int:id>')
@login_required
def item_detail(id):
    rows = sb.table('items').select(
        '*, categories(name), locations(name)'
    ).eq('id', id).limit(1).execute().data
    if not rows:
        return "<h1>Artikel nicht gefunden</h1><p><a href='/'>Zurück</a></p>", 404
    item = _normalize_item(rows[0])
    mvs = sb.table('movements').select('*').eq('item_id', id).order('created_at', desc=True).limit(10).execute().data
    return render_template('item_detail.html', item=item, movements=mvs)

# ─── Registration ─────────────────────────────────────────────────────────────

@app.route('/api/register', methods=['POST'])
@limiter.limit("10 per hour")
def register_organization():
    try:
        data = request.json
        reg_type = data.get('type', 'organization')

        if reg_type == 'organization':
            for field in ['org_name', 'email', 'username', 'password']:
                if not data.get(field):
                    return jsonify({'success': False, 'message': f'Feld "{field}" ist erforderlich'}), 400

            org_name = sanitize_string(data['org_name'], 100)
            org_slug = create_slug(org_name)
            email = data['email'].strip().lower()
            username = sanitize_string(data['username'], 50)

            is_valid, msg = validate_password(data['password'])
            if not is_valid:
                return jsonify({'success': False, 'message': msg}), 400

            # Check slug uniqueness
            existing = sb.table('organizations').select('id').eq('slug', org_slug).execute().data
            if existing:
                return jsonify({'success': False, 'message': 'Eine Organisation mit diesem Namen existiert bereits'}), 400

            # Create Supabase Auth user
            try:
                auth_resp = sb.auth.admin.create_user({
                    "email": email,
                    "password": data['password'],
                    "email_confirm": True
                })
                auth_user_id = str(auth_resp.user.id)
            except Exception as e:
                return jsonify({'success': False, 'message': f'Fehler beim Erstellen des Benutzers: {str(e)}'}), 400

            # Create organization
            org_row = sb.table('organizations').insert({
                'name': org_name,
                'slug': org_slug,
                'email': email,
                'phone': '',
                'plan': 'free',
                'max_users': 5,
                'max_items': 1000,
                'is_active': True
            }).execute().data[0]
            organization_id = org_row['id']

            # Create user row
            sb.table('users').insert({
                'auth_id': auth_user_id,
                'organization_id': organization_id,
                'username': username,
                'email': email,
                'first_name': '',
                'last_name': '',
                'is_admin': True,
                'is_org_owner': True
            }).execute()

            return jsonify({
                'success': True,
                'message': 'Organisation erfolgreich erstellt',
                'organization_id': organization_id
            })

        elif reg_type == 'user':
            for field in ['invitation_token', 'email', 'username', 'password']:
                if not data.get(field):
                    return jsonify({'success': False, 'message': f'Feld "{field}" ist erforderlich'}), 400

            token_rows = sb.table('invitation_tokens').select(
                '*, organizations(name, max_users, is_active)'
            ).eq('token', data['invitation_token']).limit(1).execute().data

            if not token_rows:
                return jsonify({'success': False, 'message': 'Ungültiger Einladungscode'}), 400

            token = token_rows[0]
            if token['is_used']:
                return jsonify({'success': False, 'message': 'Dieser Einladungscode wurde bereits verwendet'}), 400

            if token.get('expires_at'):
                expires_at = datetime.fromisoformat(token['expires_at'].replace('Z', '+00:00'))
                if datetime.now().astimezone() > expires_at:
                    return jsonify({'success': False, 'message': 'Dieser Einladungscode ist abgelaufen'}), 400

            organization_id = token['organization_id']
            org = token.get('organizations') or {}
            max_users = org.get('max_users', 5)

            user_count = sb.table('users').select('id', count='exact').eq('organization_id', organization_id).execute().count
            if user_count >= max_users:
                return jsonify({'success': False, 'message': 'Maximale Benutzeranzahl erreicht'}), 400

            email = data['email'].strip().lower()
            username = sanitize_string(data['username'], 50)

            is_valid, msg = validate_password(data['password'])
            if not is_valid:
                return jsonify({'success': False, 'message': msg}), 400

            try:
                auth_resp = sb.auth.admin.create_user({
                    "email": email,
                    "password": data['password'],
                    "email_confirm": True
                })
                auth_user_id = str(auth_resp.user.id)
            except Exception as e:
                return jsonify({'success': False, 'message': f'Fehler: {str(e)}'}), 400

            user_row = sb.table('users').insert({
                'auth_id': auth_user_id,
                'organization_id': organization_id,
                'username': username,
                'email': email,
                'first_name': '',
                'last_name': '',
                'is_admin': False,
                'is_org_owner': False
            }).execute().data[0]

            sb.table('invitation_tokens').update({
                'is_used': True,
                'used_by': user_row['id'],
                'used_at': datetime.now().isoformat()
            }).eq('id', token['id']).execute()

            return jsonify({'success': True, 'message': 'Benutzer erfolgreich registriert'})

        else:
            return jsonify({'success': False, 'message': 'Ungültiger Registrierungstyp'}), 400

    except Exception as e:
        return jsonify({'success': False, 'message': f'Fehler: {str(e)}'}), 500

# ─── API helpers ──────────────────────────────────────────────────────────────

@app.route('/api/csrf-token')
@login_required
def get_csrf_token():
    from flask_wtf.csrf import generate_csrf
    return jsonify({'csrf_token': generate_csrf()})

# ─── Dashboard ────────────────────────────────────────────────────────────────

@app.route('/api/dashboard')
@login_required
def dashboard():
    org_id = session.get('organization_id')

    total_items = sb.table('items').select('id', count='exact').eq('organization_id', org_id).execute().count
    total_categories = sb.table('categories').select('id', count='exact').eq('organization_id', org_id).execute().count
    total_locations = sb.table('locations').select('id', count='exact').eq('organization_id', org_id).execute().count

    all_items = sb.table('items').select('quantity, min_quantity, price').eq('organization_id', org_id).execute().data
    low_stock = sum(1 for i in all_items if i['quantity'] <= i['min_quantity'])
    total_value = sum((i['quantity'] or 0) * (i['price'] or 0) for i in all_items)

    return jsonify({
        'total_items': total_items,
        'total_categories': total_categories,
        'total_locations': total_locations,
        'low_stock_items': low_stock,
        'total_value': round(total_value, 2)
    })

@app.route('/api/organization-info')
@login_required
def organization_info():
    org_id = session.get('organization_id')
    rows = sb.table('organizations').select('*').eq('id', org_id).limit(1).execute().data
    if not rows:
        return jsonify({'success': False, 'message': 'Organisation nicht gefunden'}), 404

    org = rows[0]
    user_count = sb.table('users').select('id', count='exact').eq('organization_id', org_id).execute().count
    item_count = sb.table('items').select('id', count='exact').eq('organization_id', org_id).execute().count

    return jsonify({
        **org,
        'user_count': user_count,
        'item_count': item_count
    })

# ─── Invitations ──────────────────────────────────────────────────────────────

@app.route('/api/invitations', methods=['GET', 'POST'])
@login_required
@csrf_protect_api()
def invitations():
    org_id = session.get('organization_id')

    if request.method == 'POST':
        if not session.get('is_admin'):
            return jsonify({'success': False, 'message': 'Admin-Rechte erforderlich'}), 403

        token = secrets.token_urlsafe(12)
        expires_days = request.json.get('expires_days', 7)
        expires_at = (datetime.now() + timedelta(days=expires_days)).isoformat()

        row = sb.table('invitation_tokens').insert({
            'organization_id': org_id,
            'token': token,
            'created_by': session.get('user_id'),
            'expires_at': expires_at,
            'is_used': False
        }).execute().data[0]

        return jsonify({'success': True, 'token': token, 'id': row['id'], 'expires_at': expires_at})

    rows = sb.table('invitation_tokens').select('*').eq('organization_id', org_id).order('created_at', desc=True).execute().data
    return jsonify(rows)

@app.route('/api/invitations/<int:token_id>', methods=['DELETE'])
@admin_required
@csrf_protect_api()
def delete_invitation(token_id):
    org_id = session.get('organization_id')
    sb.table('invitation_tokens').delete().eq('id', token_id).eq('organization_id', org_id).execute()
    return jsonify({'success': True, 'message': 'Einladung gelöscht'})

@app.route('/api/invitations/validate', methods=['POST'])
def validate_invitation():
    token_val = request.json.get('token', '')
    rows = sb.table('invitation_tokens').select(
        '*, organizations(name)'
    ).eq('token', token_val).eq('is_used', False).limit(1).execute().data

    if not rows:
        return jsonify({'valid': False})

    token = rows[0]
    if token.get('expires_at'):
        expires_at = datetime.fromisoformat(token['expires_at'].replace('Z', '+00:00'))
        if datetime.now().astimezone() > expires_at:
            return jsonify({'valid': False})

    org = token.get('organizations') or {}
    return jsonify({'valid': True, 'organization_name': org.get('name', '')})

# ─── Users ────────────────────────────────────────────────────────────────────

@app.route('/api/users', methods=['GET', 'POST'])
@admin_required
@csrf_protect_api()
def users():
    org_id = session.get('organization_id')

    if request.method == 'POST':
        data = request.json
        email = data.get('email', '').strip().lower()
        username = sanitize_string(data.get('username'), 50)
        is_valid, msg = validate_password(data.get('password', ''))
        if not is_valid:
            return jsonify({'success': False, 'message': msg}), 400

        try:
            auth_resp = sb.auth.admin.create_user({
                "email": email,
                "password": data['password'],
                "email_confirm": True
            })
            auth_user_id = str(auth_resp.user.id)
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 400

        row = sb.table('users').insert({
            'auth_id': auth_user_id,
            'organization_id': org_id,
            'username': username,
            'email': email,
            'is_admin': bool(data.get('is_admin', False)),
            'is_org_owner': False
        }).execute().data[0]

        return jsonify({'success': True, 'id': row['id']})

    rows = sb.table('users').select('*').eq('organization_id', org_id).order('username').execute().data
    return jsonify(rows)

@app.route('/api/users/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@admin_required
@csrf_protect_api()
def user_detail(id):
    org_id = session.get('organization_id')
    rows = sb.table('users').select('*').eq('id', id).eq('organization_id', org_id).limit(1).execute().data
    if not rows:
        return jsonify({'success': False, 'message': 'Benutzer nicht gefunden'}), 404
    user = rows[0]

    if request.method == 'DELETE':
        if user.get('is_org_owner'):
            return jsonify({'success': False, 'message': 'Organisations-Eigentümer kann nicht gelöscht werden'}), 400
        if user.get('auth_id'):
            try:
                sb.auth.admin.delete_user(user['auth_id'])
            except Exception:
                pass
        sb.table('users').delete().eq('id', id).execute()
        return jsonify({'success': True, 'message': 'Benutzer gelöscht'})

    if request.method == 'PUT':
        data = request.json
        update_data = {
            'username': sanitize_string(data.get('username', user['username']), 50),
            'first_name': sanitize_string(data.get('first_name', ''), 100),
            'last_name': sanitize_string(data.get('last_name', ''), 100),
            'is_admin': bool(data.get('is_admin', False))
        }
        sb.table('users').update(update_data).eq('id', id).execute()
        return jsonify({'success': True, 'message': 'Benutzer aktualisiert'})

    return jsonify(user)

@app.route('/api/profile/change-password', methods=['POST'])
@login_required
@csrf_protect_api()
def change_user_password():
    data = request.json
    new_password = data.get('new_password', '')
    is_valid, msg = validate_password(new_password)
    if not is_valid:
        return jsonify({'success': False, 'message': msg}), 400

    auth_id = session.get('auth_id')
    try:
        sb.auth.admin.update_user_by_id(auth_id, {"password": new_password})
        return jsonify({'success': True, 'message': 'Passwort erfolgreich geändert'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

@app.route('/api/profile/update-notifications', methods=['POST'])
@login_required
@csrf_protect_api()
def update_notification_settings():
    data = request.json
    sb.table('users').update({
        'notify_low_stock': bool(data.get('notify_low_stock', True)),
        'notify_maintenance': bool(data.get('notify_maintenance', True))
    }).eq('id', session.get('user_id')).execute()
    return jsonify({'success': True, 'message': 'Einstellungen gespeichert'})

# ─── Categories ───────────────────────────────────────────────────────────────

@app.route('/api/categories', methods=['GET', 'POST'])
@login_required
@csrf_protect_api()
def categories():
    org_id = session.get('organization_id')

    if request.method == 'POST':
        data = request.json
        name = sanitize_string(data.get('name'), 100)
        if not name:
            return jsonify({'success': False, 'message': 'Name ist erforderlich'}), 400
        try:
            row = sb.table('categories').insert({
                'organization_id': org_id,
                'name': name,
                'description': sanitize_string(data.get('description', ''), 500)
            }).execute().data[0]
            return jsonify({'success': True, 'message': 'Kategorie erstellt', 'id': row['id']})
        except Exception as e:
            if 'unique' in str(e).lower():
                return jsonify({'success': False, 'message': 'Kategorie existiert bereits'}), 400
            return jsonify({'success': False, 'message': str(e)}), 400

    rows = sb.table('categories').select('*').eq('organization_id', org_id).order('name').execute().data
    return jsonify(rows)

@app.route('/api/categories/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
@csrf_protect_api()
def category(id):
    org_id = session.get('organization_id')
    rows = sb.table('categories').select('*').eq('id', id).eq('organization_id', org_id).limit(1).execute().data
    if not rows:
        return jsonify({'success': False, 'message': 'Kategorie nicht gefunden'}), 404

    if request.method == 'DELETE':
        sb.table('categories').delete().eq('id', id).execute()
        return jsonify({'success': True, 'message': 'Kategorie gelöscht'})

    if request.method == 'PUT':
        data = request.json
        sb.table('categories').update({
            'name': sanitize_string(data.get('name'), 100),
            'description': sanitize_string(data.get('description', ''), 500)
        }).eq('id', id).execute()
        return jsonify({'success': True, 'message': 'Kategorie aktualisiert'})

    return jsonify(rows[0])

# ─── Locations ────────────────────────────────────────────────────────────────

@app.route('/api/locations', methods=['GET', 'POST'])
@login_required
@csrf_protect_api()
def locations():
    org_id = session.get('organization_id')

    if request.method == 'POST':
        data = request.json
        name = sanitize_string(data.get('name'), 100)
        if not name:
            return jsonify({'success': False, 'message': 'Name ist erforderlich'}), 400
        try:
            row = sb.table('locations').insert({
                'organization_id': org_id,
                'name': name,
                'description': sanitize_string(data.get('description', ''), 500)
            }).execute().data[0]
            return jsonify({'success': True, 'message': 'Standort erstellt', 'id': row['id']})
        except Exception as e:
            if 'unique' in str(e).lower():
                return jsonify({'success': False, 'message': 'Standort existiert bereits'}), 400
            return jsonify({'success': False, 'message': str(e)}), 400

    rows = sb.table('locations').select('*').eq('organization_id', org_id).order('name').execute().data
    return jsonify(rows)

@app.route('/api/locations/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
@csrf_protect_api()
def location(id):
    org_id = session.get('organization_id')
    rows = sb.table('locations').select('*').eq('id', id).eq('organization_id', org_id).limit(1).execute().data
    if not rows:
        return jsonify({'success': False, 'message': 'Standort nicht gefunden'}), 404

    if request.method == 'DELETE':
        sb.table('locations').delete().eq('id', id).execute()
        return jsonify({'success': True, 'message': 'Standort gelöscht'})

    if request.method == 'PUT':
        data = request.json
        sb.table('locations').update({
            'name': sanitize_string(data.get('name'), 100),
            'description': sanitize_string(data.get('description', ''), 500)
        }).eq('id', id).execute()
        return jsonify({'success': True, 'message': 'Standort aktualisiert'})

    return jsonify(rows[0])

# ─── Items ────────────────────────────────────────────────────────────────────

@app.route('/api/items', methods=['GET', 'POST'])
@login_required
@csrf_protect_api()
def items():
    org_id = session.get('organization_id')
    if not org_id:
        return jsonify({'success': False, 'message': 'Keine Organisation in Session. Bitte neu einloggen.'}), 401

    if request.method == 'POST':
        data = request.json
        name = sanitize_string(data.get('name'), 255)
        if not name:
            return jsonify({'success': False, 'message': 'Name ist erforderlich'}), 400

        quantity = validate_number(data.get('quantity', 0), min_val=0, max_val=999999999)
        min_quantity = validate_number(data.get('min_quantity', 0), min_val=0, max_val=999999999)
        price = validate_number(data.get('price', 0.0), min_val=0)

        if quantity is None or min_quantity is None or price is None:
            return jsonify({'success': False, 'message': 'Ungültige numerische Werte'}), 400

        sku = sanitize_string(data.get('sku'), 100) or None
        barcode_val = sanitize_string(data.get('barcode'), 100) or None
        is_group = bool(data.get('is_group', False))
        group_id = int(data['group_id']) if data.get('group_id') else None

        if is_group and group_id:
            return jsonify({'success': False, 'message': 'Gruppen können nicht verschachtelt werden'}), 400

        if group_id:
            grp_rows = sb.table('items').select('id, is_group, organization_id').eq('id', group_id).limit(1).execute().data
            if not grp_rows or not grp_rows[0]['is_group'] or grp_rows[0]['organization_id'] != org_id:
                return jsonify({'success': False, 'message': 'Ungültige Gruppe'}), 400

        insert_data = {
            'organization_id': org_id,
            'sku': sku,
            'name': name,
            'barcode': barcode_val,
            'description': sanitize_string(data.get('description'), 1000),
            'category_id': data.get('category_id') or None,
            'location_id': data.get('location_id') or None,
            'quantity': int(quantity),
            'min_quantity': int(min_quantity),
            'unit': sanitize_string(data.get('unit', 'Stück'), 50),
            'price': float(price),
            'supplier': sanitize_string(data.get('supplier'), 255),
            'notes': sanitize_string(data.get('notes'), 1000),
            'requires_maintenance': bool(data.get('requires_maintenance', False)),
            'maintenance_interval_days': int(data['maintenance_interval_days']) if data.get('maintenance_interval_days') else None,
            'last_maintenance_date': data.get('last_maintenance_date') or None,
            'next_maintenance_date': data.get('next_maintenance_date') or None,
            'maintenance_notes': data.get('maintenance_notes') or None,
            'is_group': is_group,
            'group_id': group_id
        }

        try:
            row = sb.table('items').insert(insert_data).execute().data[0]
            return jsonify({'success': True, 'message': 'Artikel erstellt', 'id': row['id']})
        except Exception as e:
            if 'unique' in str(e).lower():
                return jsonify({'success': False, 'message': 'SKU existiert bereits'}), 400
            return jsonify({'success': False, 'message': str(e)}), 400

    # GET
    search = request.args.get('search', '')
    category = request.args.get('category')
    location_filter = request.args.get('location')
    low_stock = request.args.get('low_stock')
    group_filter = request.args.get('group')

    query = sb.table('items').select('*, categories(name), locations(name)').eq('organization_id', org_id)

    if category:
        query = query.eq('category_id', int(category))
    if location_filter:
        query = query.eq('location_id', int(location_filter))
    if group_filter:
        query = query.eq('group_id', int(group_filter))

    rows = query.order('name').execute().data

    # Python-side filters
    if search:
        s = search.lower()
        rows = [r for r in rows if s in (r.get('name') or '').lower()
                or s in (r.get('sku') or '').lower()
                or s in (r.get('description') or '').lower()]
    if low_stock:
        rows = [r for r in rows if r['quantity'] <= r['min_quantity']]

    # Add group names via Python lookup
    group_ids = {r['group_id'] for r in rows if r.get('group_id')}
    group_map = {}
    if group_ids:
        grp_rows = sb.table('items').select('id, name').in_('id', list(group_ids)).execute().data
        group_map = {g['id']: g['name'] for g in grp_rows}

    result = []
    for r in rows:
        r = _normalize_item(r)
        r['group_name'] = group_map.get(r.get('group_id'))
        result.append(r)

    # Sort groups first
    result.sort(key=lambda x: (not x.get('is_group'), x.get('name', '')))
    return jsonify(result)

@app.route('/api/items/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
@csrf_protect_api()
def item(id):
    try:
        org_id = session.get('organization_id')

        if request.method == 'DELETE':
            children = sb.table('items').select('id', count='exact').eq('group_id', id).eq('organization_id', org_id).execute().count
            if children > 0:
                return jsonify({'success': False, 'message': f'Gruppe enthält noch {children} Artikel'}), 400
            sb.table('items').delete().eq('id', id).eq('organization_id', org_id).execute()
            return jsonify({'success': True, 'message': 'Artikel gelöscht'})

        if request.method == 'PUT':
            data = request.json
            barcode_val = data.get('barcode') or None
            is_group = bool(data.get('is_group', False))
            group_id = int(data['group_id']) if data.get('group_id') else None

            if is_group and group_id:
                return jsonify({'success': False, 'message': 'Gruppen können nicht verschachtelt werden'}), 400

            if group_id:
                grp_rows = sb.table('items').select('id, is_group, organization_id').eq('id', group_id).limit(1).execute().data
                if not grp_rows or not grp_rows[0]['is_group'] or grp_rows[0]['organization_id'] != org_id:
                    return jsonify({'success': False, 'message': 'Ungültige Gruppe'}), 400

            if not is_group:
                children = sb.table('items').select('id', count='exact').eq('group_id', id).eq('organization_id', org_id).execute().count
                if children > 0:
                    return jsonify({'success': False, 'message': f'Diese Gruppe enthält noch {children} Artikel'}), 400

            sb.table('items').update({
                'sku': data.get('sku') or None,
                'name': data['name'],
                'barcode': barcode_val,
                'description': data.get('description'),
                'category_id': data.get('category_id') or None,
                'location_id': data.get('location_id') or None,
                'quantity': data.get('quantity', 0),
                'min_quantity': data.get('min_quantity', 0),
                'unit': data.get('unit', 'Stück'),
                'price': data.get('price', 0.0),
                'supplier': data.get('supplier'),
                'notes': data.get('notes'),
                'requires_maintenance': bool(data.get('requires_maintenance', False)),
                'maintenance_interval_days': data.get('maintenance_interval_days') or None,
                'last_maintenance_date': data.get('last_maintenance_date') or None,
                'next_maintenance_date': data.get('next_maintenance_date') or None,
                'maintenance_notes': data.get('maintenance_notes'),
                'is_group': is_group,
                'group_id': group_id,
                'updated_at': datetime.now().isoformat()
            }).eq('id', id).eq('organization_id', org_id).execute()
            return jsonify({'success': True, 'message': 'Artikel aktualisiert'})

        rows = sb.table('items').select('*, categories(name), locations(name)').eq('id', id).eq('organization_id', org_id).limit(1).execute().data
        if not rows:
            return jsonify({}), 404
        item_data = _normalize_item(rows[0])
        if item_data.get('group_id'):
            grp_rows = sb.table('items').select('name').eq('id', item_data['group_id']).limit(1).execute().data
            item_data['group_name'] = grp_rows[0]['name'] if grp_rows else None
        return jsonify(item_data)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# ─── Movements / Withdrawal ───────────────────────────────────────────────────

@app.route('/api/items/<int:id>/move', methods=['POST'])
@login_required
@csrf_protect_api()
def move_item(id):
    data = request.json
    move_type = data['type']
    quantity = int(data['quantity'])
    org_id = session.get('organization_id')
    user_id = session.get('user_id')

    rows = sb.table('items').select('quantity').eq('id', id).limit(1).execute().data
    if not rows:
        return jsonify({'success': False, 'message': 'Artikel nicht gefunden'}), 404

    current_qty = rows[0]['quantity']

    if move_type == 'in':
        new_qty = current_qty + quantity
    else:
        if current_qty < quantity:
            return jsonify({'success': False, 'message': 'Nicht genügend Bestand'}), 400
        new_qty = current_qty - quantity

    sb.table('items').update({'quantity': new_qty, 'updated_at': datetime.now().isoformat()}).eq('id', id).execute()

    slip_id = None
    if move_type == 'out':
        slip_row = sb.table('withdrawal_slips').insert({
            'organization_id': org_id,
            'user_id': user_id,
            'reference': data.get('reference'),
            'notes': data.get('notes')
        }).execute().data[0]
        slip_id = slip_row['id']

    sb.table('movements').insert({
        'item_id': id,
        'user_id': user_id,
        'type': move_type,
        'quantity': quantity,
        'reference': data.get('reference'),
        'notes': data.get('notes'),
        'slip_id': slip_id
    }).execute()

    return jsonify({'success': True, 'message': 'Bewegung gebucht', 'new_quantity': new_qty, 'slip_id': slip_id})

@app.route('/api/withdrawals/batch', methods=['POST'])
@login_required
@csrf_protect_api()
def create_batch_withdrawal():
    data = request.json
    items_list = data.get('items', [])
    org_id = session.get('organization_id')
    user_id = session.get('user_id')

    if not items_list:
        return jsonify({'success': False, 'message': 'Keine Artikel angegeben'}), 400

    for entry in items_list:
        rows = sb.table('items').select('quantity').eq('id', int(entry['item_id'])).eq('organization_id', org_id).limit(1).execute().data
        if not rows:
            return jsonify({'success': False, 'message': f'Artikel {entry["item_id"]} nicht gefunden'}), 404
        if rows[0]['quantity'] < int(entry['quantity']):
            return jsonify({'success': False, 'message': f'Nicht genügend Bestand für Artikel {entry["item_id"]}'}), 400

    slip_row = sb.table('withdrawal_slips').insert({
        'organization_id': org_id,
        'user_id': user_id,
        'reference': data.get('reference'),
        'notes': data.get('notes')
    }).execute().data[0]
    slip_id = slip_row['id']

    for entry in items_list:
        item_id = int(entry['item_id'])
        quantity = int(entry['quantity'])
        rows = sb.table('items').select('quantity').eq('id', item_id).limit(1).execute().data
        current_qty = rows[0]['quantity']
        sb.table('items').update({'quantity': current_qty - quantity, 'updated_at': datetime.now().isoformat()}).eq('id', item_id).execute()
        sb.table('movements').insert({
            'item_id': item_id,
            'user_id': user_id,
            'type': 'out',
            'quantity': quantity,
            'reference': data.get('reference'),
            'notes': entry.get('notes'),
            'slip_id': slip_id
        }).execute()

    return jsonify({'success': True, 'slip_id': slip_id})

@app.route('/slip/<int:slip_id>')
@login_required
def print_slip(slip_id):
    org_id = session.get('organization_id')

    slip_rows = sb.table('withdrawal_slips').select('*, users(first_name, last_name, username), organizations(name)').eq('id', slip_id).eq('organization_id', org_id).limit(1).execute().data
    if not slip_rows:
        return 'Entnahmeschein nicht gefunden', 404

    slip = slip_rows[0]
    u = slip.pop('users') or {}
    o = slip.pop('organizations') or {}
    slip['user_name'] = f"{u.get('first_name','')} {u.get('last_name','')}".strip() or u.get('username', '')
    slip['org_name'] = o.get('name', '')

    mv_rows = sb.table('movements').select(
        'quantity, notes, items(name, sku, unit, locations(name))'
    ).eq('slip_id', slip_id).order('id').execute().data

    movements = []
    for m in mv_rows:
        itm = m.pop('items') or {}
        loc = itm.pop('locations') or {}
        movements.append({
            'quantity': m['quantity'],
            'notes': m.get('notes'),
            'item_name': itm.get('name'),
            'sku': itm.get('sku'),
            'unit': itm.get('unit'),
            'location_name': loc.get('name')
        })

    return render_template('entnahmeschein.html', slip=slip, movements=movements)

@app.route('/api/items/<int:id>/movements')
@login_required
def item_movements(id):
    rows = sb.table('movements').select('*').eq('item_id', id).order('created_at', desc=True).execute().data
    return jsonify(rows)

# ─── QR / Barcode routes ──────────────────────────────────────────────────────

@app.route('/api/items/<int:id>/qrcode')
@login_required
def get_item_qrcode(id):
    rows = sb.table('items').select('*').eq('id', id).limit(1).execute().data
    if not rows:
        return jsonify({'success': False, 'message': 'Artikel nicht gefunden'}), 404
    return send_file(generate_qr_code(id, rows[0]), mimetype='image/png')

@app.route('/api/items/<int:id>/qrcode-base64')
@login_required
def get_item_qrcode_base64(id):
    rows = sb.table('items').select('*').eq('id', id).limit(1).execute().data
    if not rows:
        return jsonify({'success': False, 'message': 'Artikel nicht gefunden'}), 404
    return jsonify({'success': True, 'qrcode': generate_qr_code_base64(id, rows[0])})

@app.route('/api/items/<int:id>/barcode')
@login_required
def get_item_barcode(id):
    rows = sb.table('items').select('*').eq('id', id).limit(1).execute().data
    if not rows:
        return jsonify({'success': False, 'message': 'Artikel nicht gefunden'}), 404
    return send_file(generate_barcode(id, rows[0]), mimetype='image/png')

@app.route('/api/items/<int:id>/barcode-base64')
@login_required
def get_item_barcode_base64(id):
    rows = sb.table('items').select('*').eq('id', id).limit(1).execute().data
    if not rows:
        return jsonify({'success': False, 'message': 'Artikel nicht gefunden'}), 404
    return jsonify({'success': True, 'barcode': generate_barcode_base64(id, rows[0])})

@app.route('/api/items/search-barcode', methods=['GET'])
@login_required
def search_item_by_barcode():
    barcode_val = request.args.get('barcode', '').strip()
    org_id = session.get('organization_id')
    if not barcode_val:
        return jsonify({'success': False, 'message': 'Kein Barcode angegeben'}), 400

    if barcode_val.startswith('ITEM'):
        try:
            item_id = int(barcode_val.replace('ITEM', '').lstrip('0') or '0')
            rows = sb.table('items').select('*').eq('id', item_id).eq('organization_id', org_id).limit(1).execute().data
            if rows:
                return jsonify({'success': True, 'item': rows[0]})
        except Exception:
            pass

    rows = sb.table('items').select('*').eq('barcode', barcode_val).eq('organization_id', org_id).limit(1).execute().data
    if rows:
        return jsonify({'success': True, 'item': rows[0]})

    return jsonify({'success': False, 'message': 'Artikel nicht gefunden', 'barcode': barcode_val, 'suggest_create': True}), 404

# ─── Image upload ─────────────────────────────────────────────────────────────

@app.route('/api/items/<int:id>/upload-image', methods=['POST'])
@login_required
@csrf_protect_api()
def upload_item_image(id):
    org_id = session.get('organization_id')
    rows = sb.table('items').select('*').eq('id', id).eq('organization_id', org_id).limit(1).execute().data
    if not rows:
        return jsonify({'success': False, 'message': 'Artikel nicht gefunden'}), 404

    if 'image' not in request.files:
        return jsonify({'success': False, 'message': 'Keine Datei hochgeladen'}), 400

    file = request.files['image']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'Keine Datei ausgewählt'}), 400
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'message': f'Ungültiger Dateityp'}), 400

    # Delete old image
    old_path = rows[0].get('image_path')
    if old_path:
        old_file = os.path.join(UPLOAD_FOLDER, old_path)
        if os.path.exists(old_file):
            os.remove(old_file)

    filename = secure_filename_custom(file.filename)
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    sb.table('items').update({'image_path': filename, 'updated_at': datetime.now().isoformat()}).eq('id', id).execute()
    return jsonify({'success': True, 'message': 'Bild hochgeladen', 'image_path': filename})

@app.route('/api/items/<int:id>/delete-image', methods=['DELETE'])
@login_required
@csrf_protect_api()
def delete_item_image(id):
    org_id = session.get('organization_id')
    rows = sb.table('items').select('image_path').eq('id', id).eq('organization_id', org_id).limit(1).execute().data
    if not rows:
        return jsonify({'success': False, 'message': 'Artikel nicht gefunden'}), 404

    image_path = rows[0].get('image_path')
    if not image_path:
        return jsonify({'success': False, 'message': 'Kein Bild vorhanden'}), 404

    filepath = os.path.join(UPLOAD_FOLDER, image_path)
    if os.path.exists(filepath):
        os.remove(filepath)

    sb.table('items').update({'image_path': None, 'updated_at': datetime.now().isoformat()}).eq('id', id).execute()
    return jsonify({'success': True, 'message': 'Bild gelöscht'})

# ─── Export ───────────────────────────────────────────────────────────────────

@app.route('/api/export/csv')
@login_required
def export_csv():
    org_id = session.get('organization_id')
    rows = sb.table('items').select('*, categories(name), locations(name)').eq('organization_id', org_id).order('name').execute().data

    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(['SKU', 'Name', 'Beschreibung', 'Kategorie', 'Standort', 'Menge', 'Einheit', 'Mindestbestand', 'Preis', 'Lieferant'])
    for r in rows:
        cat = (r.get('categories') or {}).get('name', '')
        loc = (r.get('locations') or {}).get('name', '')
        writer.writerow([r.get('sku',''), r['name'], r.get('description',''), cat, loc,
                         r['quantity'], r.get('unit',''), r['min_quantity'], r.get('price',0), r.get('supplier','')])

    return Response(si.getvalue(), mimetype='text/csv',
                    headers={'Content-disposition': 'attachment; filename=inventory_export.csv'})

@app.route('/api/export/excel')
@login_required
def export_excel():
    org_id = session.get('organization_id')
    rows = sb.table('items').select('*, categories(name), locations(name)').eq('organization_id', org_id).order('name').execute().data

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventar"

    header_fill = PatternFill(start_color="7c3aed", end_color="7c3aed", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['ID', 'Name', 'SKU', 'Menge', 'Preis (€)', 'Gesamtwert (€)', 'Kategorie', 'Standort', 'Lieferant', 'Wartung', 'Nächste Wartung', 'Notizen']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row_idx, r in enumerate(rows, 2):
        cat = (r.get('categories') or {}).get('name', '')
        loc = (r.get('locations') or {}).get('name', '')
        total_value = (r.get('quantity', 0) or 0) * (r.get('price', 0) or 0)
        ws.cell(row=row_idx, column=1, value=r['id']).border = border
        ws.cell(row=row_idx, column=2, value=r['name']).border = border
        ws.cell(row=row_idx, column=3, value=r.get('sku', '')).border = border
        ws.cell(row=row_idx, column=4, value=r.get('quantity', 0)).border = border
        price_cell = ws.cell(row=row_idx, column=5, value=r.get('price', 0))
        price_cell.number_format = '#,##0.00 €'; price_cell.border = border
        total_cell = ws.cell(row=row_idx, column=6, value=total_value)
        total_cell.number_format = '#,##0.00 €'; total_cell.border = border
        ws.cell(row=row_idx, column=7, value=cat).border = border
        ws.cell(row=row_idx, column=8, value=loc).border = border
        ws.cell(row=row_idx, column=9, value=r.get('supplier', '')).border = border
        ws.cell(row=row_idx, column=10, value='Ja' if r.get('requires_maintenance') else 'Nein').border = border
        ws.cell(row=row_idx, column=11, value=r.get('next_maintenance_date', '')).border = border
        ws.cell(row=row_idx, column=12, value=r.get('notes', '')).border = border

    for col, width in enumerate([8, 30, 15, 10, 12, 15, 15, 15, 20, 10, 15, 40], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='inventory_export.xlsx')

@app.route('/api/export/pdf')
@login_required
def export_pdf():
    org_id = session.get('organization_id')
    rows = sb.table('items').select('*, categories(name), locations(name)').eq('organization_id', org_id).order('name').execute().data

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm, leftMargin=1*cm, rightMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"Inventarliste - {session.get('organization_name', '')}", styles['Title']))
    elements.append(Paragraph(f"Erstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    table_data = [['SKU', 'Name', 'Menge', 'Einheit', 'Preis', 'Kategorie', 'Standort']]
    for r in rows:
        cat = (r.get('categories') or {}).get('name', '')
        loc = (r.get('locations') or {}).get('name', '')
        table_data.append([r.get('sku',''), r['name'], str(r.get('quantity',0)), r.get('unit',''),
                           f"{r.get('price',0):.2f} €", cat, loc])

    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#7c3aed')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f3ff')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name='inventory_export.pdf')

# ─── Stats ────────────────────────────────────────────────────────────────────

@app.route('/api/stats/value-trend')
@login_required
def get_value_trend():
    org_id = session.get('organization_id')
    all_items = sb.table('items').select('quantity, price').eq('organization_id', org_id).execute().data
    total_value = round(sum((i.get('quantity',0) or 0) * (i.get('price',0) or 0) for i in all_items), 2)
    data = []
    for i in range(30, -1, -1):
        d = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        data.append({'date': d, 'value': total_value})
    return jsonify(data)

@app.route('/api/stats/category-distribution')
@login_required
def get_category_distribution():
    org_id = session.get('organization_id')
    rows = sb.table('items').select('category_id, quantity, price, categories(name)').eq('organization_id', org_id).execute().data

    buckets = {}
    for r in rows:
        cat = (r.get('categories') or {}).get('name') or 'Ohne Kategorie'
        if cat not in buckets:
            buckets[cat] = {'category': cat, 'count': 0, 'value': 0}
        buckets[cat]['count'] += 1
        buckets[cat]['value'] += (r.get('quantity',0) or 0) * (r.get('price',0) or 0)

    return jsonify(sorted(buckets.values(), key=lambda x: x['count'], reverse=True))

@app.route('/api/stats/top-items')
@login_required
def get_top_items():
    org_id = session.get('organization_id')
    rows = sb.table('items').select('id, name, sku, quantity, price, categories(name), locations(name)').eq('organization_id', org_id).execute().data
    for r in rows:
        r['category_name'] = (r.pop('categories') or {}).get('name')
        r['location_name'] = (r.pop('locations') or {}).get('name')
        r['total_value'] = (r.get('quantity',0) or 0) * (r.get('price',0) or 0)
    rows.sort(key=lambda x: x['total_value'], reverse=True)
    return jsonify(rows[:10])

@app.route('/api/stats/recent-activity')
@login_required
def get_recent_activity():
    org_id = session.get('organization_id')
    rows = sb.table('items').select('id, name, quantity, categories(name)').eq('organization_id', org_id).order('updated_at', desc=True).limit(10).execute().data
    result = []
    for r in rows:
        cat = (r.pop('categories') or {}).get('name', '')
        result.append({'id': r['id'], 'name': r['name'], 'quantity': r['quantity'],
                       'category_name': cat, 'action': 'Aktualisiert',
                       'timestamp': datetime.now().isoformat()})
    return jsonify(result)

# ─── Print QR/Barcode all ─────────────────────────────────────────────────────

@app.route('/api/items/qrcodes/print')
@login_required
def print_all_qrcodes():
    org_id = session.get('organization_id')
    rows = sb.table('items').select('*').eq('organization_id', org_id).order('name').execute().data
    items_with_codes = []
    for r in rows:
        r['qrcode'] = generate_qr_code_base64(r['id'], r)
        items_with_codes.append(r)
    return render_template('print_qrcodes.html', items=items_with_codes)

@app.route('/api/items/barcodes/print')
@login_required
def print_all_barcodes():
    org_id = session.get('organization_id')
    rows = sb.table('items').select('*').eq('organization_id', org_id).order('name').execute().data
    items_with_codes = []
    for r in rows:
        r['barcode'] = generate_barcode_base64(r['id'], r)
        items_with_codes.append(r)
    return render_template('print_barcodes.html', items=items_with_codes)

# ─── Maintenance Types ────────────────────────────────────────────────────────

@app.route('/api/maintenance/types', methods=['GET', 'POST'])
@login_required
def maintenance_types_api():
    org_id = session.get('organization_id')

    if request.method == 'POST':
        data = request.json
        name = sanitize_string(data.get('name'), 100)
        if not name:
            return jsonify({'success': False, 'message': 'Name ist erforderlich'}), 400
        try:
            row = sb.table('maintenance_types').insert({
                'organization_id': org_id,
                'name': name,
                'description': sanitize_string(data.get('description'), 500),
                'color': data.get('color', '#7c3aed'),
                'icon': data.get('icon', 'wrench'),
                'default_interval_days': data.get('default_interval_days')
            }).execute().data[0]
            return jsonify({'success': True, 'message': 'Wartungstyp erstellt', 'id': row['id']})
        except Exception as e:
            if 'unique' in str(e).lower():
                return jsonify({'success': False, 'message': 'Ein Wartungstyp mit diesem Namen existiert bereits'}), 400
            return jsonify({'success': False, 'message': str(e)}), 400

    rows = sb.table('maintenance_types').select('*').eq('organization_id', org_id).eq('is_active', True).order('name').execute().data
    return jsonify(rows)

@app.route('/api/maintenance/types/<int:type_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
@csrf_protect_api()
def maintenance_type_detail(type_id):
    org_id = session.get('organization_id')
    rows = sb.table('maintenance_types').select('*').eq('id', type_id).eq('organization_id', org_id).limit(1).execute().data
    if not rows:
        return jsonify({'success': False, 'message': 'Wartungstyp nicht gefunden'}), 404
    mt = rows[0]

    if request.method == 'DELETE':
        sb.table('maintenance_types').update({'is_active': False}).eq('id', type_id).execute()
        return jsonify({'success': True, 'message': 'Wartungstyp deaktiviert'})

    if request.method == 'PUT':
        data = request.json
        sb.table('maintenance_types').update({
            'name': sanitize_string(data.get('name'), 100),
            'description': sanitize_string(data.get('description'), 500),
            'color': data.get('color', mt['color']),
            'icon': data.get('icon', mt['icon']),
            'default_interval_days': data.get('default_interval_days')
        }).eq('id', type_id).eq('organization_id', org_id).execute()
        return jsonify({'success': True, 'message': 'Wartungstyp aktualisiert'})

    return jsonify(mt)

# ─── Maintenance Schedules ────────────────────────────────────────────────────

@app.route('/api/maintenance/schedules/<int:item_id>', methods=['GET', 'POST'])
@login_required
def item_schedules_api(item_id):
    org_id = session.get('organization_id')
    item_rows = sb.table('items').select('id').eq('id', item_id).eq('organization_id', org_id).limit(1).execute().data
    if not item_rows:
        return jsonify({'success': False, 'message': 'Artikel nicht gefunden'}), 404

    if request.method == 'POST':
        data = request.json
        type_id = data.get('maintenance_type_id')
        interval_days = data.get('interval_days')
        if not type_id or not interval_days:
            return jsonify({'success': False, 'message': 'Wartungstyp und Intervall sind erforderlich'}), 400

        next_date = data.get('next_date')
        if not next_date:
            next_date = (datetime.now() + timedelta(days=int(interval_days))).strftime('%Y-%m-%d')

        try:
            row = sb.table('item_maintenance_schedules').insert({
                'item_id': item_id,
                'maintenance_type_id': int(type_id),
                'interval_days': int(interval_days),
                'last_date': data.get('last_date'),
                'next_date': next_date,
                'assigned_user_id': data.get('assigned_user_id'),
                'priority': data.get('priority', 'normal'),
                'notes': sanitize_string(data.get('notes'), 500),
                'is_active': True
            }).execute().data[0]
            return jsonify({'success': True, 'message': 'Wartungsplan erstellt', 'id': row['id']})
        except Exception as e:
            if 'unique' in str(e).lower():
                return jsonify({'success': False, 'message': 'Wartungsplan für diesen Typ existiert bereits'}), 400
            return jsonify({'success': False, 'message': str(e)}), 400

    rows = sb.table('item_maintenance_schedules').select(
        '*, maintenance_types(name, color, icon), users(username)'
    ).eq('item_id', item_id).eq('is_active', True).execute().data

    result = []
    for r in rows:
        mt = r.pop('maintenance_types') or {}
        u = r.pop('users') or {}
        r['type_name'] = mt.get('name')
        r['type_color'] = mt.get('color')
        r['type_icon'] = mt.get('icon')
        r['assigned_username'] = u.get('username')
        result.append(r)
    return jsonify(result)

@app.route('/api/maintenance/schedules/entry/<int:schedule_id>', methods=['PUT', 'DELETE'])
@login_required
@csrf_protect_api()
def schedule_entry_api(schedule_id):
    org_id = session.get('organization_id')

    if request.method == 'DELETE':
        sb.table('item_maintenance_schedules').update({'is_active': False}).eq('id', schedule_id).execute()
        return jsonify({'success': True, 'message': 'Wartungsplan deaktiviert'})

    data = request.json
    sb.table('item_maintenance_schedules').update({
        'interval_days': data.get('interval_days'),
        'next_date': data.get('next_date'),
        'priority': data.get('priority', 'normal'),
        'notes': sanitize_string(data.get('notes'), 500),
        'updated_at': datetime.now().isoformat()
    }).eq('id', schedule_id).execute()
    return jsonify({'success': True, 'message': 'Wartungsplan aktualisiert'})

@app.route('/api/maintenance/schedules/entry/<int:schedule_id>/assign', methods=['PUT'])
@login_required
@csrf_protect_api()
def assign_schedule(schedule_id):
    data = request.json
    sb.table('item_maintenance_schedules').update({
        'assigned_user_id': data.get('assigned_user_id'),
        'updated_at': datetime.now().isoformat()
    }).eq('id', schedule_id).execute()
    return jsonify({'success': True, 'message': 'Benutzer zugewiesen'})

# ─── Maintenance Due ──────────────────────────────────────────────────────────

@app.route('/api/maintenance/due')
@login_required
def get_due_maintenance():
    org_id = session.get('organization_id')
    warning_days = request.args.get('warning_days', 30, type=int)
    type_id_filter = request.args.get('type_id', type=int)
    assigned_user_id_filter = request.args.get('assigned_user_id', type=int)
    status_filter = request.args.get('status')

    today = datetime.now().date()
    warning_date = (today + timedelta(days=warning_days)).isoformat()

    query = sb.table('item_maintenance_schedules').select(
        '*, items(id, name, sku, image_path, organization_id, categories(name), locations(name)), maintenance_types(id, name, color, icon), users(username)'
    ).eq('is_active', True).lte('next_date', warning_date).not_.is_('next_date', 'null')

    if type_id_filter:
        query = query.eq('maintenance_type_id', type_id_filter)
    if assigned_user_id_filter:
        query = query.eq('assigned_user_id', assigned_user_id_filter)

    rows = query.order('next_date').execute().data

    result = []
    for r in rows:
        itm = r.pop('items') or {}
        if itm.get('organization_id') != org_id:
            continue
        mt = r.pop('maintenance_types') or {}
        u = r.pop('users') or {}
        cat = (itm.pop('categories', None) or {}).get('name')
        loc = (itm.pop('locations', None) or {}).get('name')

        next_d = date.fromisoformat(r['next_date']) if r.get('next_date') else None
        days_until = (next_d - today).days if next_d else None
        if next_d and next_d <= today:
            mstatus = 'overdue'
        elif next_d and next_d <= today + timedelta(days=7):
            mstatus = 'due_soon'
        else:
            mstatus = 'ok'

        entry = {
            **r,
            'schedule_id': r['id'],
            'item_id': itm.get('id'),
            'item_name': itm.get('name'),
            'sku': itm.get('sku'),
            'image_path': itm.get('image_path'),
            'type_id': mt.get('id'),
            'type_name': mt.get('name'),
            'type_color': mt.get('color'),
            'type_icon': mt.get('icon'),
            'category_name': cat,
            'location_name': loc,
            'assigned_username': u.get('username'),
            'maintenance_status': mstatus,
            'days_until_maintenance': days_until
        }
        result.append(entry)

    if status_filter and status_filter != 'all':
        result = [r for r in result if r['maintenance_status'] == status_filter]

    return jsonify(result)

@app.route('/api/maintenance/my-tasks')
@login_required
def get_my_maintenance_tasks():
    org_id = session.get('organization_id')
    user_id = session.get('user_id')
    today = datetime.now().date()

    rows = sb.table('item_maintenance_schedules').select(
        '*, items(id, name, sku, organization_id, categories(name), locations(name)), maintenance_types(name, color, icon)'
    ).eq('assigned_user_id', user_id).eq('is_active', True).not_.is_('next_date', 'null').order('next_date').execute().data

    result = []
    for r in rows:
        itm = r.pop('items') or {}
        if itm.get('organization_id') != org_id:
            continue
        mt = r.pop('maintenance_types') or {}
        cat = (itm.pop('categories', None) or {}).get('name')
        loc = (itm.pop('locations', None) or {}).get('name')
        next_d = date.fromisoformat(r['next_date']) if r.get('next_date') else None
        days_until = (next_d - today).days if next_d else None
        mstatus = 'overdue' if (next_d and next_d <= today) else 'due_soon'
        result.append({
            **r, 'schedule_id': r['id'],
            'item_id': itm.get('id'), 'item_name': itm.get('name'), 'sku': itm.get('sku'),
            'type_name': mt.get('name'), 'type_color': mt.get('color'), 'type_icon': mt.get('icon'),
            'category_name': cat, 'location_name': loc,
            'maintenance_status': mstatus, 'days_until_maintenance': days_until
        })
    return jsonify(result)

@app.route('/api/maintenance/calendar')
@login_required
def get_maintenance_calendar():
    org_id = session.get('organization_id')
    start = request.args.get('start', datetime.now().strftime('%Y-%m-01'))
    end_default = (datetime.now().replace(day=1) + timedelta(days=32)).replace(day=1).strftime('%Y-%m-%d')
    end = request.args.get('end', end_default)
    today_str = datetime.now().strftime('%Y-%m-%d')

    rows = sb.table('item_maintenance_schedules').select(
        'id, next_date, priority, items(id, name, organization_id), maintenance_types(name, color)'
    ).eq('is_active', True).gte('next_date', start).lte('next_date', end).order('next_date').execute().data

    result = []
    for r in rows:
        itm = r.pop('items') or {}
        if itm.get('organization_id') != org_id:
            continue
        mt = r.pop('maintenance_types') or {}
        result.append({
            'schedule_id': r['id'],
            'next_date': r['next_date'],
            'priority': r.get('priority'),
            'item_id': itm.get('id'),
            'item_name': itm.get('name'),
            'type_name': mt.get('name'),
            'type_color': mt.get('color'),
            'status': 'overdue' if r['next_date'] <= today_str else 'upcoming'
        })
    return jsonify(result)

# ─── Maintenance Complete ─────────────────────────────────────────────────────

@app.route('/api/maintenance/complete/<int:schedule_id>', methods=['POST'])
@login_required
@csrf_protect_api()
def complete_maintenance(schedule_id):
    org_id = session.get('organization_id')
    data = request.json

    sched_rows = sb.table('item_maintenance_schedules').select(
        '*, items(id, organization_id), maintenance_types(name)'
    ).eq('id', schedule_id).limit(1).execute().data

    if not sched_rows:
        return jsonify({'success': False, 'message': 'Wartungsplan nicht gefunden'}), 404

    schedule = sched_rows[0]
    itm = schedule.get('items') or {}
    if itm.get('organization_id') != org_id:
        return jsonify({'success': False, 'message': 'Wartungsplan nicht gefunden'}), 404

    maintenance_date = data.get('maintenance_date', datetime.now().strftime('%Y-%m-%d'))
    maintenance_dt = datetime.strptime(maintenance_date, '%Y-%m-%d')
    next_maintenance = (maintenance_dt + timedelta(days=schedule['interval_days'])).strftime('%Y-%m-%d')

    sb.table('item_maintenance_schedules').update({
        'last_date': maintenance_date,
        'next_date': next_maintenance,
        'updated_at': datetime.now().isoformat()
    }).eq('id', schedule_id).execute()

    sb.table('items').update({
        'last_maintenance_date': maintenance_date,
        'next_maintenance_date': next_maintenance,
        'updated_at': datetime.now().isoformat()
    }).eq('id', itm['id']).execute()

    cost_parts = float(data.get('cost_parts', 0) or 0)
    cost_labor = float(data.get('cost_labor', 0) or 0)
    cost = cost_parts + cost_labor

    history_row = sb.table('maintenance_history').insert({
        'item_id': itm['id'],
        'user_id': session.get('user_id'),
        'organization_id': org_id,
        'maintenance_type_id': schedule['maintenance_type_id'],
        'schedule_id': schedule_id,
        'maintenance_date': maintenance_date,
        'performed_by': session.get('username'),
        'notes': data.get('notes', ''),
        'next_maintenance_date': next_maintenance,
        'cost': cost,
        'cost_parts': cost_parts,
        'cost_labor': cost_labor,
        'cost_notes': sanitize_string(data.get('cost_notes'), 500)
    }).execute().data[0]
    history_id = history_row['id']

    for result in data.get('checklist_results', []):
        sb.table('maintenance_checklist_results').insert({
            'maintenance_history_id': history_id,
            'checklist_item_id': result.get('checklist_item_id'),
            'is_checked': bool(result.get('is_checked', False)),
            'notes': sanitize_string(result.get('notes'), 500)
        }).execute()

    return jsonify({
        'success': True,
        'message': 'Wartung erfolgreich abgeschlossen',
        'next_maintenance_date': next_maintenance,
        'history_id': history_id
    })

# ─── Maintenance History ──────────────────────────────────────────────────────

@app.route('/api/maintenance/history/<int:item_id>')
@login_required
def get_maintenance_history(item_id):
    org_id = session.get('organization_id')
    rows = sb.table('maintenance_history').select(
        '*, maintenance_types(name, color)'
    ).eq('item_id', item_id).eq('organization_id', org_id).order('maintenance_date', desc=True).execute().data

    result = []
    for h in rows:
        mt = h.pop('maintenance_types') or {}
        h['type_name'] = mt.get('name')
        h['type_color'] = mt.get('color')
        checklist = sb.table('maintenance_checklist_results').select(
            '*, checklist_items(description, is_required)'
        ).eq('maintenance_history_id', h['id']).execute().data
        for c in checklist:
            ci = c.pop('checklist_items') or {}
            c['item_description'] = ci.get('description')
            c['is_required'] = ci.get('is_required')
        h['checklist_results'] = checklist
        result.append(h)
    return jsonify(result)

# ─── Maintenance Checklists ───────────────────────────────────────────────────

@app.route('/api/maintenance/checklists/<int:type_id>', methods=['GET', 'POST'])
@login_required
def maintenance_checklists_api(type_id):
    org_id = session.get('organization_id')
    mt_rows = sb.table('maintenance_types').select('id').eq('id', type_id).eq('organization_id', org_id).limit(1).execute().data
    if not mt_rows:
        return jsonify({'success': False, 'message': 'Wartungstyp nicht gefunden'}), 404

    if request.method == 'POST':
        data = request.json
        checklist_name = sanitize_string(data.get('name', 'Standard-Checkliste'), 200)
        items_data = data.get('items', [])

        existing = sb.table('maintenance_checklists').select('id').eq('maintenance_type_id', type_id).eq('organization_id', org_id).limit(1).execute().data
        if existing:
            checklist_id = existing[0]['id']
            sb.table('maintenance_checklists').update({'name': checklist_name}).eq('id', checklist_id).execute()
            sb.table('checklist_items').delete().eq('checklist_id', checklist_id).execute()
        else:
            cl_row = sb.table('maintenance_checklists').insert({
                'maintenance_type_id': type_id,
                'organization_id': org_id,
                'name': checklist_name
            }).execute().data[0]
            checklist_id = cl_row['id']

        for idx, ci in enumerate(items_data):
            sb.table('checklist_items').insert({
                'checklist_id': checklist_id,
                'description': sanitize_string(ci.get('description'), 500),
                'sort_order': idx,
                'is_required': bool(ci.get('is_required', False))
            }).execute()

        return jsonify({'success': True, 'message': 'Checkliste gespeichert', 'id': checklist_id})

    cl_rows = sb.table('maintenance_checklists').select('id, name').eq('maintenance_type_id', type_id).eq('organization_id', org_id).limit(1).execute().data
    if not cl_rows:
        return jsonify({'id': None, 'name': None, 'items': []})

    cl = cl_rows[0]
    ci_rows = sb.table('checklist_items').select('*').eq('checklist_id', cl['id']).order('sort_order').execute().data
    return jsonify({'id': cl['id'], 'name': cl['name'], 'items': ci_rows})

@app.route('/api/maintenance/checklist-results/<int:history_id>')
@login_required
def get_checklist_results(history_id):
    rows = sb.table('maintenance_checklist_results').select(
        '*, checklist_items(description, is_required)'
    ).eq('maintenance_history_id', history_id).execute().data
    for r in rows:
        ci = r.pop('checklist_items') or {}
        r['item_description'] = ci.get('description')
        r['is_required'] = ci.get('is_required')
    return jsonify(rows)

# ─── Maintenance Stats ────────────────────────────────────────────────────────

@app.route('/api/maintenance/stats')
@login_required
def get_maintenance_stats():
    org_id = session.get('organization_id')
    today = datetime.now().date()
    seven_days = today + timedelta(days=7)
    month_start = today.replace(day=1).isoformat()

    sched_rows = sb.table('item_maintenance_schedules').select(
        'next_date, items(organization_id)'
    ).eq('is_active', True).not_.is_('next_date', 'null').execute().data

    overdue = due_soon = ok_count = 0
    for r in sched_rows:
        if (r.get('items') or {}).get('organization_id') != org_id:
            continue
        nd = date.fromisoformat(r['next_date'])
        if nd <= today:
            overdue += 1
        elif nd <= seven_days:
            due_soon += 1
        else:
            ok_count += 1

    hist_rows = sb.table('maintenance_history').select('cost, maintenance_date').eq('organization_id', org_id).gte('maintenance_date', month_start).execute().data
    monthly_cost = sum(h.get('cost', 0) or 0 for h in hist_rows)

    ninety_days_ago = (today - timedelta(days=90)).isoformat()
    total_completed = sb.table('maintenance_history').select('id', count='exact').eq('organization_id', org_id).gte('maintenance_date', ninety_days_ago).execute().count

    return jsonify({
        'overdue': overdue,
        'due_soon': due_soon,
        'ok': ok_count,
        'monthly_cost': round(monthly_cost, 2),
        'compliance_rate': 100.0,
        'total_completed_90d': total_completed
    })

@app.route('/api/maintenance/cost-report')
@login_required
def get_maintenance_cost_report():
    org_id = session.get('organization_id')
    period = int(request.args.get('period', 12))
    group_by = request.args.get('group_by', 'month')
    months_ago = (datetime.now() - timedelta(days=period * 30)).strftime('%Y-%m-%d')

    rows = sb.table('maintenance_history').select(
        '*, maintenance_types(name), items(name)'
    ).eq('organization_id', org_id).gte('maintenance_date', months_ago).order('maintenance_date').execute().data

    buckets = {}
    for r in rows:
        mt = (r.get('maintenance_types') or {}).get('name', 'Unbekannt')
        itm = (r.get('items') or {}).get('name', 'Unbekannt')
        date_str = r.get('maintenance_date', '')

        if group_by == 'month':
            label = date_str[:7] if date_str else 'Unbekannt'
        elif group_by == 'type':
            label = mt
        else:
            label = itm

        if label not in buckets:
            buckets[label] = {'label': label, 'total_cost': 0, 'parts': 0, 'labor': 0, 'count': 0}
        buckets[label]['total_cost'] += r.get('cost', 0) or 0
        buckets[label]['parts'] += r.get('cost_parts', 0) or 0
        buckets[label]['labor'] += r.get('cost_labor', 0) or 0
        buckets[label]['count'] += 1

    return jsonify(sorted(buckets.values(), key=lambda x: x['label']))

@app.route('/api/maintenance/compliance-report')
@login_required
def get_maintenance_compliance_report():
    org_id = session.get('organization_id')
    rows = sb.table('maintenance_history').select(
        'maintenance_date, next_maintenance_date, maintenance_types(name)'
    ).eq('organization_id', org_id).limit(100).execute().data
    return jsonify({'records': rows, 'compliance_rate': 100.0})

@app.route('/api/maintenance/export/csv')
@login_required
def export_maintenance_csv():
    org_id = session.get('organization_id')
    rows = sb.table('maintenance_history').select(
        '*, items(name, sku), maintenance_types(name)'
    ).eq('organization_id', org_id).order('maintenance_date', desc=True).execute().data

    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(['Datum', 'Artikel', 'SKU', 'Wartungstyp', 'Durchgeführt von', 'Kosten', 'Notizen'])
    for r in rows:
        itm = r.get('items') or {}
        mt = r.get('maintenance_types') or {}
        writer.writerow([r.get('maintenance_date',''), itm.get('name',''), itm.get('sku',''),
                         mt.get('name',''), r.get('performed_by',''),
                         r.get('cost',0), r.get('notes','')])

    return Response(si.getvalue(), mimetype='text/csv',
                    headers={'Content-disposition': 'attachment; filename=maintenance_export.csv'})

@app.route('/api/maintenance/export/pdf')
@login_required
def export_maintenance_pdf():
    org_id = session.get('organization_id')
    rows = sb.table('maintenance_history').select(
        '*, items(name, sku), maintenance_types(name)'
    ).eq('organization_id', org_id).order('maintenance_date', desc=True).limit(100).execute().data

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"Wartungshistorie - {session.get('organization_name', '')}", styles['Title']),
        Spacer(1, 0.5*cm)
    ]

    table_data = [['Datum', 'Artikel', 'Wartungstyp', 'Durchgeführt von', 'Kosten']]
    for r in rows:
        itm = r.get('items') or {}
        mt = r.get('maintenance_types') or {}
        table_data.append([r.get('maintenance_date',''), itm.get('name',''),
                           mt.get('name',''), r.get('performed_by',''),
                           f"{r.get('cost',0):.2f} €"])

    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#7c3aed')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f3ff')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name='maintenance_export.pdf')

# ─── Label Templates ──────────────────────────────────────────────────────────

@app.route('/api/label-templates', methods=['GET', 'POST'])
@login_required
def label_templates():
    org_id = session.get('organization_id')

    if request.method == 'POST':
        data = request.json
        try:
            row = sb.table('label_templates').insert({
                'organization_id': org_id,
                'name': data['name'],
                'description': data.get('description', ''),
                'width_mm': data['width_mm'],
                'height_mm': data['height_mm'],
                'layout_config': data['layout_config']
            }).execute().data[0]
            return jsonify({'success': True, 'message': 'Template gespeichert', 'id': row['id']})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 400

    rows = sb.table('label_templates').select('*').eq('organization_id', org_id).order('name').execute().data
    return jsonify(rows)

@app.route('/api/label-templates/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
@csrf_protect_api()
def label_template(id):
    org_id = session.get('organization_id')
    rows = sb.table('label_templates').select('*').eq('id', id).eq('organization_id', org_id).limit(1).execute().data
    if not rows:
        return jsonify({'success': False, 'message': 'Template nicht gefunden'}), 404

    if request.method == 'DELETE':
        sb.table('label_templates').delete().eq('id', id).execute()
        return jsonify({'success': True, 'message': 'Template gelöscht'})

    if request.method == 'PUT':
        data = request.json
        sb.table('label_templates').update({
            'name': data['name'],
            'description': data.get('description', ''),
            'width_mm': data['width_mm'],
            'height_mm': data['height_mm'],
            'layout_config': data['layout_config'],
            'is_default': bool(data.get('is_default', False)),
            'updated_at': datetime.now().isoformat()
        }).eq('id', id).execute()
        return jsonify({'success': True, 'message': 'Template aktualisiert'})

    return jsonify(rows[0])

@app.route('/api/items/print-custom-labels')
@login_required
def print_custom_labels():
    org_id = session.get('organization_id')
    template_id = request.args.get('template_id', type=int)
    item_id = request.args.get('item_id', type=int)
    category = request.args.get('category', type=int)
    location = request.args.get('location', type=int)

    tmpl_rows = sb.table('label_templates').select('*').eq('id', template_id).eq('organization_id', org_id).limit(1).execute().data
    if not tmpl_rows:
        return 'Template nicht gefunden', 404

    query = sb.table('items').select('*, categories(name), locations(name)').eq('organization_id', org_id)
    if item_id:
        query = query.eq('id', item_id)
    elif category:
        query = query.eq('category_id', category)
    elif location:
        query = query.eq('location_id', location)

    rows = query.order('name').execute().data
    items_with_codes = []
    for r in rows:
        r = _normalize_item(r)
        r['barcode'] = generate_barcode_base64(r['id'], r)
        r['qrcode'] = generate_qr_code_base64(r['id'], r)
        items_with_codes.append(r)

    return render_template('print_custom_labels.html', items=items_with_codes, template=tmpl_rows[0])

# ─── Misc ─────────────────────────────────────────────────────────────────────

@app.route('/debug/session')
def debug_session():
    return jsonify({
        'logged_in': session.get('logged_in'),
        'user_id': session.get('user_id'),
        'username': session.get('username'),
        'organization_id': session.get('organization_id'),
        'organization_name': session.get('organization_name'),
        'is_admin': session.get('is_admin'),
        'is_org_owner': session.get('is_org_owner')
    })

# ─── Startup ──────────────────────────────────────────────────────────────────

def check_system_status():
    class C:
        G = '\033[92m'; W = '\033[93m'; F = '\033[91m'; E = '\033[0m'; B = '\033[1m'; C2 = '\033[96m'

    print(f"\n{C.C2}{C.B}    StockMaster - Supabase Edition{C.E}")
    print(f"{C.B}    {'-' * 50}{C.E}\n")

    checks = []
    try:
        count = sb.table('organizations').select('id', count='exact').execute().count
        checks.append(("Supabase DB", f"Connected ({count} orgs)", "green"))
    except Exception as e:
        checks.append(("Supabase DB", f"ERROR: {str(e)[:40]}", "red"))

    checks.append(("Upload Folder", "Ready" if os.path.exists(UPLOAD_FOLDER) else "Missing", "green" if os.path.exists(UPLOAD_FOLDER) else "yellow"))
    checks.append(("Templates", "OK", "green" if os.path.exists('templates/index.html') else "red"))
    checks.append(("Static JS", "OK", "green" if os.path.exists('static/app.js') else "red"))

    for name, status, color in checks:
        sym = "+" if color == "green" else ("!" if color == "yellow" else "X")
        col = C.G if color == "green" else (C.W if color == "yellow" else C.F)
        print(f"    {col}{sym}{C.E} {C.B}{name:<20}{C.E} {status}")

    print(f"\n{C.B}    URL: http://localhost:5000{C.E}\n")

if __name__ == '__main__':
    check_system_status()
    app.run(
        host=os.getenv('HOST', '0.0.0.0'),
        port=int(os.getenv('PORT', 5000)),
        debug=os.getenv('DEBUG', 'False') == 'True'
    )
